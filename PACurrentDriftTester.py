"""
2023-08-30
@author Jin Her
"""
from dataclasses import dataclass
import schedule
import time
from openpyxl import Workbook
from Multibranch.Test_Suite import *


@dataclass
class Results:
    current: float
    temperature: float
    local_temperature: float


class PADriftResultCollector:
    def __init__(self, radio, branch: int):
        self.radio = radio
        self.branch = branch

        self.time = 0
        self.results = None

    def collect_current(self):
        return self.radio.read_current(self.branch)

    def collect_temperature(self):
        return self.radio.read_temperature(self.branch)
    
    def collect_local_temperature(self):
        return self.radio.read_local_temperature(self.branch)

    def get_results(self):
        self.result = Results(current=self.collect_current(), temperature=self.collect_temperature(), local_temperature=self.collect_temperature()) 
        return self.results


class PADriftResultWriter:
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active

        self.row = 1
        self.column = 1

        current_lbl = self.ws.cell(row=self.row, column=self.column)
        temp_lbl = self.ws.cell(row=self.row, column=self.column+1)
        local_temp_lbl = self.ws.cell(row=self.row, column=self.column+2)

        current_lbl.value = 'Current'
        temp_lbl.value = 'Temperature'
        local_temp_lbl.value = 'Local Temperature'

    def get_row(self):
        self.row += 1
        return self.row

    def write(self, results: Results):
        self.row = self.get_row()

        current = self.ws.cell(row=self.row, column=self.column)
        temperature = self.ws.cell(row=self.row, column=self.column+1)
        local_temperature = self.ws.cell(row=self.row, column=self.column+2)

        current.value = results.current
        temperature.value = results.temperature
        local_temperature.value = results.local_temperature

    def save(self, stem_of_file_name: str):
        self.wb.save(f"{stem_of_file_name}.xlsx")


if __name__ == '__main__':
    branch = 1
    max_time = 43200 # 12 hours
    start_time = time.time()
    stem_of_file_name = f"Branch{branch}_{max_time}seconds"

    radio = Radio()
    radio.connect_BBB()
    
    collector = PADriftResultCollector(radio, branch)
    writer = PADriftResultWriter()

    schedule.every(30).minutes.do(writer.write, collector.get_results())

    while (time.time() - start_time) < max_time:
        schedule.run_pending()
        time.sleep(1)

    writer.save(stem_of_file_name)
    radio.disconnect_BBB()

